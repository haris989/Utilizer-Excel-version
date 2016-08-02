import matplotlib.pyplot as plt
import matplotlib.patches as mpatch
from openpyxl import Workbook
from openpyxl import load_workbook
from fpdf import FPDF
import os
print("Please wait while your shipment's picksheet is being generated...")
pdf = FPDF()
k_plan_workbook = Workbook()
k_plan_workbook = load_workbook('excel.xlsm', read_only=True)
dump_sheet=k_plan_workbook.active
xlimat=110
ylimat=101
rowno=0
coors=[[]]
for row in dump_sheet.rows:

    for cell in row:
       coors[rowno].append(cell.value)
    coors.append([])
    rowno=rowno+1
layerval=[]
#print(coors)
for i in range(1,len(coors)-1):
    if(coors[i][4]!=coors[i-1][4]):
        layerval.append([coors[i][6], coors[i][4]])
#print(coors[0][0])
#layerval=[rollcage,layery,]
#print(layerval)

layer=0
layer_max=len(layerval)
imagelist=[]
while (layer<layer_max):


    fig, ax = plt.subplots()
    rectangles={}
    for i in range (0,(len(coors)-2)):
        if (coors[i+1][4]==layerval[layer][1] and coors[i+1][6]==layerval[layer][0]):
            rectangles.update({str(coors[i+1][7]) + str(i+1) : mpatch.Rectangle((int(coors[i+1][3]),int(coors[i+1][5])), int(coors[i+1][0]), int(coors[i+1][1]))})



    for r in rectangles:
        ax.add_artist(rectangles[r])
        rx, ry = rectangles[r].get_xy()
        cx = rx + rectangles[r].get_width()/2.0
        cy = ry + rectangles[r].get_height()/2.0

        ax.annotate(r, (cx, cy), color='w', weight='bold',
                    fontsize=9, ha='center', va='center')

    ax.set_xlim((0, xlimat))
    ax.set_ylim((0, ylimat))
    ax.set_aspect('equal')
    plt.xlabel('Roll Cage Length')
    plt.ylabel('Roll Cage Breadth')

    plt.title('Roll Cage '+str(layerval[layer][0])+' - Layer' +str(layer+1))



    fig.savefig('Roll Cage '+str(layerval[layer][0])+' - Layer ' +str(layer+1)+ '.png')
    imagelist.append('Roll Cage '+str(layerval[layer][0])+' - Layer ' +str(layer+1)+ '.png')
    layer=layer+1
    #plt.show()
    plt.close(fig)

i=0

while(i<len(imagelist)):
    pdf.add_page()
    pdf.image(imagelist[i],0,30,70,70)
    if(i+1<len(imagelist)):
        pdf.image(imagelist[i+1],70,30,70,70)
    if(i+2<len(imagelist)):
        pdf.image(imagelist[i+2],140,30,70,70)

    if(i+3<len(imagelist)):
        pdf.image(imagelist[i+3],0,110,70,70)
    if(i+4<len(imagelist)):
        pdf.image(imagelist[i+4],70,110,70,70)
    if(i+5<len(imagelist)):
        pdf.image(imagelist[i+5],140,110,70,70)

    if(i+6<len(imagelist)):
        pdf.image(imagelist[i+6],0,190,70,70)
    if(i+7<len(imagelist)):
        pdf.image(imagelist[i+7],70,190,70,70)
    if(i+8<len(imagelist)):
        pdf.image(imagelist[i+8],140,190,70,70)
    i=i+9
pdf.output("PickSheet.pdf", "F")

for file in imagelist:
    os.remove(file)